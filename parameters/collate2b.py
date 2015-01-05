#!/usr/bin/python
# -*- coding: utf-8 -*-
import sqlite3
import os
import numpy as np
import pandas as pd
import datetime
import pytz
import math
import osgeo.ogr
import shutil
import openpyxl

original = os.path.join('data', 'processed', 'laketemp.sqlite')
working = os.path.join('data', 'processed', 'laketemp_working.sqlite')
shutil.copyfile(original, working)

conn = sqlite3.connect(working)

### NVE
conn.execute('insert into providers (provider_name) values (?)', ('NVE', ))
c = conn.execute('select provider_id from providers where provider_name = ?', 
                 ('NVE', ))
pr_id = c.fetchall()[0][0]

### NVE Ånund data time series
p = os.path.join('data', 'NVE data', 'Hessen vanntemploggere.xlsx')
wb = openpyxl.load_workbook(p)
for ws in wb.worksheets:
    print('reading sheet %s' % ws.title)
    print('first 5 rows:')
    for r in ws.rows[:5]:
        print('    %s %s' % (r[0].value, r[1].value))
    name0 = ws['B2'].value
    name = '_'.join(name0.split(' ')[2:])
    name = name.replace(u'Å', u'A').replace(u'å', u'a')
    name = name.replace(u'Ø', u'O').replace(u'ø', u'o')
    name = name.replace(u'Æ', u'AE').replace(u'æ', u'ae')
    name9 = name.lower()[:9]
    print('uses this name for identification: %s' % name9)
    subdf = namesdf[namesdf['name'] == name9]
    print(subdf.to_string())
    this_lnr = dict1[ws.title]
    if this_lnr is None:
        print('** PLACE UNKNOWN TODO')
    else:
        print('** chosen the following vatnlnr: %s' % this_lnr)
        print(namesdf[namesdf['vatnlnr'] == this_lnr].to_string())
        ebint = eblnrdf[eblnrdf['vatn_lnr'] == this_lnr]['ebint'].tolist()[0]
        print('** EBint is %s' % ebint)
        conn.execute('insert into lakes (name, eb_int, provider_id) values (?, ?, ?)',
                     (name0, ebint, pr_id))
        for r in ws.rows[5:]:
            day, month, year = map(int, r[0].value.split()[0].split('.'))
            val = r[1].value
            conn.execute('insert into temperature values (?, ?, ?, ?, ?, ?)', 
                         (pr_id, ebint, 0, datetime.date(year, month, day), 0, val))     
    print('')

### NVE Ånund data depth profile
p = os.path.join('data', 'NVE data', 'Hessen vanntempvertikal.xlsx')
wb = openpyxl.load_workbook(p)
nsheets = len(wb.worksheets)
sheetnames = ['hessen%i' % num for num in range(1, 1 + nsheets)]

# for sheetname in sheetnames:
#     ws = wb.get_sheet_by_name(sheetname)
#     # print('first 4 rows:')
#     # for r in ws.rows[:4]:
#     #     print('    %s %s' % (r[0].value, r[1].value))
#     name00 = ws['B2'].value
#     name0 = name00.split('NV')[0].split(' v/')[0].split(',')[0].lstrip()
#     name = name0
#     name = name.replace(u'Å', u'A').replace(u'å', u'a')
#     name = name.replace(u'Ø', u'O').replace(u'ø', u'o')
#     name = name.replace(u'Æ', u'AE').replace(u'æ', u'ae')
#     name9 = name.lower()[:9]
#     # print('uses this name for identification: %s' % name9)
#     subdf = namesdf[namesdf['name'] == name9]
#     if subdf.shape[0] > 1:
#         print('unidentified sheet %s' % ws.title)
#         print(name00)
#         print(subdf.to_string())
#         print('')

# print('------')
# print('These are not found in fish database')
# for sheetname in sheetnames:
#     ws = wb.get_sheet_by_name(sheetname)
#     # print('first 4 rows:')
#     # for r in ws.rows[:4]:
#     #     print('    %s %s' % (r[0].value, r[1].value))
#     name00 = ws['B2'].value
#     name0 = name00.split('NV')[0].split(' v/')[0].split(',')[0].lstrip()
#     name = name0
#     name = name.replace(u'Å', u'A').replace(u'å', u'a')
#     name = name.replace(u'Ø', u'O').replace(u'ø', u'o')
#     name = name.replace(u'Æ', u'AE').replace(u'æ', u'ae')
#     name9 = name.lower()[:9]
#     # print('uses this name for identification: %s' % name9)
#     subdf = namesdf[namesdf['name'] == name9]
#     if subdf.shape[0] == 0:
#         print('unidentified sheet %s %s' % (ws.title, name00))

dict2 = {'hessen4': 125, 
         'hessen15': 274, 
         'hessen17': 149, 
         'hessen30': 515, 
         'hessen43': 1058,
         ## hessen46
         'hessen53': 1649, 
         'hessen57': 1806, 
         'hessen61': 910, 
         'hessen64': 2376, 
         'hessen69': 2243, 
         ## 'hessen72': 
         'hessen6': 291, 
         'hessen16': 220, 
         'hessen21': 522,
         ## 'hessen22': hessen22 Steinsfjorden v/Herøya
         'hessen23': 542, ## Eikeren v/Gunhildrud  fish database id wrong? 542 includes the skumvannet (541)
         'hessen24': 541, ## again fish database id wrong? 
         'hessen36': 6, 
         'hessen44': 1087, 
         'hessen51': 1702, 
         'hessen68': 2399, 
         ## 'hessen71': 2193 ?? Joatkajavre v/Årdoaivi 
         'hessen73': 2279}
         
for sheetname in sheetnames:
    ws = wb.get_sheet_by_name(sheetname)
    print('first 4 rows:')
    for r in ws.rows[:4]:
        print('    %s %s' % (r[0].value, r[1].value))
    name00 = ws['B2'].value
    name0 = name00.split('NV')[0].split(' v/')[0].split(',')[0].lstrip()
    name = name0
    name = name.replace(u'Å', u'A').replace(u'å', u'a')
    name = name.replace(u'Ø', u'O').replace(u'ø', u'o')
    name = name.replace(u'Æ', u'AE').replace(u'æ', u'ae')
    name9 = name.lower()[:9]
    print('uses this name for identification: %s' % name9)
    subdf = namesdf[namesdf['name'] == name9]
    if subdf.shape[0] == 1:
        this_vatnlnr = subdf['vatnlnr'].tolist()[0]
        print(name00)
        print(subdf.to_string())
        print('** identified as vatnlnr: %s' % this_vatnlnr)
    elif sheetname in dict2.keys():
        this_vatnlnr = dict2[sheetname]
        print('** identified as vatnlnr: %s' % this_vatnlnr)
    else:
        print('** for the moment pass this one TODO')
        print('')
        continue
    ebint = eblnrdf[eblnrdf['vatn_lnr'] == this_vatnlnr]['ebint'].tolist()[0]
    print('** EBint is %s' % ebint)
    print('')
    c = conn.execute('select station_id from lakes where eb_int = ? and provider_id = ?', 
                     (ebint, pr_id))
    station_id = len(c.fetchall()) + 1
    conn.execute('insert into lakes (name, eb_int, provider_id, station_id) values (?, ?, ?, ?)',
                 (name0, ebint, pr_id, station_id))
    for r in ws.rows[4:]:
        A = r[0].value
        B = r[1].value
        if type(A) == datetime.datetime:
            date = A.date() ## ignoring time
            continue
        else:
            conn.execute('insert into temperature values (?, ?, ?, ?, ?, ?)', 
                         (pr_id, ebint, station_id, date, -1 * A, B))  

conn.commit()
conn.close()
